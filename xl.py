import xlrd as xread
import xlwt as xwrite
from openpyxl import Workbook, load_workbook
import os

def main():
    list_xls_file = os.listdir("./xls")
    list_xls_file = [i for i in list_xls_file if i[-4:] != '.log']

    list_russia_file = os.listdir("./russia2")
    list_russia_file = [i for i in list_russia_file if i[-4:] != '.log']

    res = []

    for file in list_xls_file:
        rb = xread.open_workbook(f'./xls/{file}')
        sheet = rb.sheet_by_index(0)
        for rownum in range(sheet.nrows):
            row = sheet.row_values(rownum)
            res_row = []
            for c_el in row:
                res_row.append(c_el)
            res_row.append(file.replace('.xls','').replace('HHCompany',''))
            res.append(res_row)
    
    print(len(res))

    for file in list_russia_file:
        rb = xread.open_workbook(f'./russia2/{file}')
        sheet = rb.sheet_by_index(1)
        for rownum in range(sheet.nrows):
            row = sheet.row_values(rownum)
            res_row = []
            for c_el in row:
                res_row.append(c_el)
            res_row.append('Россия')
            res.append(res_row)
    
    print(len(res))

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title='HHCompany')
    # ws = wb.add_sheet('HHCompany')

    for row in res:
        ws.append(row)
    
    wb.save('HHCompany.xlsx')
    

if __name__ == "__main__":
    main()